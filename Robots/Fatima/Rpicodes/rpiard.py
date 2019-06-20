
import mazelogic # se encarga de la logica en el modo maze(opcion de debugger usando procesing)
import RPi.GPIO as GPIO
import sys
import serial
import time
import os
import socket
import bluetooth
import select
from thread import *
import threading
import math
from openpyxl import load_workbook
from Queue import Queue
GPIO.setmode (GPIO.BCM) #nomenclatura GPIO# no numero de pin
ledS=19 #led  de estado de espera/eleccion de modo
b1=25 #boton para confirmacion de conexion con arduino/MODO MANUAL(activacion en lectura de 1)
b2=16 #boton para modo Maze(activacion en lectura de 1)
nmos=12 #activa rele para encender arduino
GPIO.setwarnings(False)
GPIO.setup(ledS,GPIO.OUT)
GPIO.setup(nmos,GPIO.OUT)
GPIO.setup(b1,GPIO.IN,pull_up_down=GPIO.PUD_DOWN) #la raspberry actuara cuando reciba 3.3v
GPIO.setup(b2,GPIO.IN,pull_up_down=GPIO.PUD_DOWN)
####INCIALIZACION DE PINES PARA L298N####
#Control Thread#
print_lock = threading.Lock()
q = Queue(maxsize=0)
#Control Thread#
motorA1=4 # in1
motorA2=23# in3
motorB1=17# in2
motorB2=22# in4
motorena=18
motorenb=26
GPIO.setup(motorA1,GPIO.OUT)
GPIO.setup(motorA2,GPIO.OUT)
GPIO.setup(motorB1,GPIO.OUT)
GPIO.setup(motorB2,GPIO.OUT)
GPIO.setup(motorena,GPIO.OUT)
GPIO.setup(motorenb,GPIO.OUT)
ena=GPIO.PWM(motorena,500)
enb=GPIO.PWM(motorenb,500)
ena.start(0)
enb.start(0)
####INCIALIZACION DE PINES PARA L298N####
##Encoder##
Ebgn=22 #Salida
Eend=21 #Entrada
GPIO.setup(Eend,GPIO.IN)
GPIO.setup(Ebgn,GPIO.OUT)
k=0 #contador dummy
##Encoder##


###CONTROL DE INCIIALIZACION###
flag=0 #activa la secuencia de activacion y acepta el modo debug
flag1=0 #se encarga de salir del while en el que se elije el modo debug
b1F=0 #elige el modo maze
b2F=0 #elige el modo manual
enaS=1 # Flag para correr codigo una sola vez
ena2S=1 #Flag para correr codigo una sola vez
debug =0 #control de modo debug
ena_Sensor=0
rst_Sensor=[0]
###CONTROL DE INCIIALIZACION###
###tcp###
imu=[0,0,0] # yaw, pitch, roll
incli=[0,0,0,0,0,0,0,0,0,0,0,0]
rad=['s',0,0] #'dire','ang','distance'
HOST= '192.168.25.110'
PORT= 6794 # Revisar contra el cliente
###tcp###
##Variables para toma de decision##
rect=[0,0] # rect_x rect_y
Dt=['Ndty',0]
Dn=[0,0,0,0]
dn=[0,0,0,0]
i=1
checker=[0]*177 #evita la eleccion de seguir recto cuando no hay nada en los alrededores
wb = load_workbook(filename = 'PROYECTO DSP DATA DE 4 CASOS (version 1).xlsx', data_only=True)
sheet_ranges = wb['Casos para algoritmo']
##Variables para toma de decision##

######Movimiento de los motores######
def detenerse():
    ena.ChangeDutyCycle(0)
    enb.ChangeDutyCycle(0)
    GPIO.output(motorA1, GPIO.LOW)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.LOW)
    GPIO.output(motorB1, GPIO.LOW)


def adelante(mode):
    
    ena.ChangeDutyCycle(80)  
    enb.ChangeDutyCycle(80)
    GPIO.output(motorA1, GPIO.LOW)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.HIGH)
    GPIO.output(motorB2, GPIO.HIGH)	
    if (mode):
        ardS.write('12')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()


def izquierda(mode):
    ena.ChangeDutyCycle(70)
    GPIO.output(motorA1, GPIO.LOW)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.HIGH)
    GPIO.output(motorB2, GPIO.LOW)
    if (mode):
        ardS.write('6')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()


def spinizq(mode):
    ena.ChangeDutyCycle(70)
    enb.ChangeDutyCycle(70)
    GPIO.output(motorA1, GPIO.LOW)
    GPIO.output(motorA2, GPIO.HIGH)
    GPIO.output(motorB1, GPIO.HIGH)
    GPIO.output(motorB2, GPIO.LOW)
    if (mode):
        ardS.write('6')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()


def derecha(mode):
    enb.ChangeDutyCycle(70)
    GPIO.output(motorA1, GPIO.LOW)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.LOW)
    GPIO.output(motorB2, GPIO.HIGH)
    if (mode):
        ardS.write('6')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()


def spinder(mode):
    enb.ChangeDutyCycle(70)
    ena.ChangeDutyCycle(70)
    GPIO.output(motorA1, GPIO.HIGH)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.LOW)
    GPIO.output(motorB2, GPIO.HIGH)
    if (mode):
        ardS.write('6')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()
		
def case4(mode):
    enb.ChangeDutyCycle(70)
    ena.ChangeDutyCycle(70)
    GPIO.output(motorA1, GPIO.HIGH)
    GPIO.output(motorA2, GPIO.LOW)
    GPIO.output(motorB1, GPIO.LOW)
    GPIO.output(motorB2, GPIO.HIGH)
    if (mode):
        ardS.write('24')
        GPIO.output(Ebng,GPIO.HIGH) #Inicia conteo
        if(GPIO.input(Eend)== 1): #retraso con polling divino
            continue
            GPIO.output(Ebgn,GPIO.LOW)
            detenerse()
######Movimiento de los motores######
######Funcion Thread######
def threaded(c): #manual mediante thread
    while True:
        #print("Thread ejectuado")
        data=c.recv(1024)
        data_decode=data.decode("utf-8")
        data_str=str(data_decode)
        if not data:
            print("Bye Bye")
            print_lock.release()
            break
        if data_str == 'w':
            adelante(0)
        elif data_str == 'q':
            spinizq(0)
        elif data_str == 'e':
            spinder(0)
        elif data_str == 's':
            detenerse()
        elif data_str == 'a':
            izquierda(0)
        elif data_str == 'd':
            derecha(0)
        else:
            detenerse()
        #print (data)
    c.close()
######Funcion Thread######

######Funcion Thread######
def threadedMaze(q): #maze control mediante thread
    while True :  #caso 1 = izquierda caso2=hacia delante caso3= derecha caso4= encerrado
        data_str=q.get()
        
        if data_str == 1 :
            print("adelante(1)")
            adelante(1)
        elif data_str == 0:
            print("spinizq(1)")
            spinizq(1)
        elif data_str == 2:
            print("spinder(1)")
            spinder(1)
        elif data_str == 3:
            print("spinder(1) encerrado")
            case4(1) 
        else:
            print("detenerse()")
            detenerse()
######Funcion Thread######

#######Envio de data usando TCP########
def txData ():
    #data = str(rad[1])+','+str(rad[2])+','+str(imu[0])+','+str(imu[1])+','+str(imu[2])+','+'s'
    data_to = str(rad[1])+','+str(rad[2])+'s'
    data_toSend=data_to.encode('utf-8')
    packet_len=len(data_toSend)
    diff=12-packet_len
    packet_adj=data_to + 's'*diff
    packet_toSend=packet_adj.encode('utf-8')
    #print(len(packet_toSend))
    try:
        conn.sendall(packet_toSend)
    except socket.error:
        print ('No se pudo enviar la informacion')
        print ("Modulos Desactivados")
        ardS.write(b'E')
        ardS.write(b's')
        conn.close()
        time.sleep(2)
        sys.exit()
########Envio de dara usanto TCP######
#arreglo para funcion rpiard
def remap(x,in_min,in_max,out_min,out_max):
    return (x-in_min)*(out_max-out_min)/(in_max-in_min)+out_min
#arreglo para funcion rpiard
####adquisicion datat#####
def rpiard(logic):
    data_raw = ardS.readline()
    if data_raw:
        datastr = data_raw.decode("utf-8")
        dsplit = datastr.split(",")
        if (dsplit[0] == 'I'):
            # print(dsplit)
            imu[0] = float(dsplit[1])
            imu[1] = float(dsplit[2])
            imu[2] = float(dsplit[3])

        if (dsplit[0] == 'R'):
            if(dsplit[1]=="CCW"):
                rad[0] = dsplit[1]
                rad[1] = float(dsplit[2])
                rad[2] = dsplit[3]
                rad[2] = rad[2].rstrip('\r\n')
                rad[2] = float(rad[2])
                rect[0]=rad[2]*math.cos(math.radians(rad[1]))
                rect[1]=rad[2]*math.sin(math.radians(rad[1]))
                ########LOGICA DE MAZE########
                if (logic):
                    with open ("radarData.csv", "a") as pos:
                        pos.write("%s, %s \n" % ( rad[1],rad[2]))
                    i= remap(rad[1],6,170,1,165)
                    
                    if (i <166):
                        #caso#1
                        e=sheet_ranges.cell(row=i,column=5).value
                        f=sheet_ranges.cell(row=i,column=6).value
                        #print("e-f= ",e,f,i,rad[1])
                        dn[0]=(((rect[0]-e)**2+(rect[1]-f)**2)**0.5)
                        Dn[0]=Dn[0]+dn[0]
                        #caso #2
                        l=sheet_ranges.cell(row=i,column=12).value
                        m=sheet_ranges.cell(row=i,column=13).value
                        #print("l-m= ",l,m,i,rad[1])
                        dn[1]=(((rect[0]-l)**2+(rect[1]-m)**2)**0.5)
                        Dn[1]=Dn[1]+dn[1]
                        #caso #3
                        s=sheet_ranges.cell(row=i,column=19).value
                        t=sheet_ranges.cell(row=i,column=20).value
                        dn[2]=(((rect[0]-s)**2+(rect[1]-t)**2)**0.5)
                        Dn[2]=Dn[2]+dn[2]
                        #print("s-t= ",s,t,i,rad[1])
                        #caso #4
                        y=sheet_ranges.cell(row=i,column=25).value
                        z=sheet_ranges.cell(row=i,column=26).value
                        dn[3]=(((rect[0]-y)**2+(rect[1]-z)**2)**0.5)
                        Dn[3]=Dn[3]+dn[3]
                        #|  print("y-z= ",y,z,i,rad[1])
                        global k
                        checker[k]=rad[1] #guardar medidas tomadas
                        k += 1
                        
                        with open ("MinimoCuadrado.csv", "a") as pos:
                            pos.write("%s, %s, %s, %s, %s, %s, %s, %s, %s \n" % (rad[1],dn[0],Dn[0],dn[1],Dn[1],dn[2],Dn[2],dn[3],Dn[3]))
                    if (i == 165):
                        print("Analisis completo")
                        Dn[0]=round(Dn[0],4)
                        Dn[1]=round(Dn[1],4)
                        Dn[2]=round(Dn[2],4)
                        Dn[3]=round(Dn[3],4)
                        check=sum(checker)
                        #Distancia Menor para determinar caso
                        if(check<100):
                            detenerse()
                            Dt[0]='Ndty'
                            Dn[0]=0
                            Dn[1]=0
                            Dn[2]=0
                            Dn[3]=0
                            i=1
                            k=0
                        else:
                            #hacer eleccion
                            Dt[0]=Dn.index(min(Dn))
                            q.put(Dt[0])
                            with open ("Decision.csv", "a") as pos:
                                pos.write("%s \n" % (Dt[0]))
                            Dt[0]='Ndty'
                            Dn[0]=0
                            Dn[1]=0
                            Dn[2]=0
                            Dn[3]=0
                            i=1
                            k=0
                    
                    if(debug):
                        txData()
                    
                if(logic==0):
                    txData()
                    with open ("reg0-180.csv", "a") as pos:
                        pos.write("%s, %s \n" % ( rad[1],rad[2]))
                        
####adquisiscion de data####
print("###incio del programa###")
GPIO.output(nmos,GPIO.LOW) #apaga rele
time.sleep(2)
GPIO.output(nmos,GPIO.HIGH) #enciende rele
ardS = serial.Serial("/dev/serial0", baudrate = 115200) # en espera de level shifter
#ardS = serial.Serial("/dev/ttyUSB0", baudrate =115200)

try:
    #Prende Led de estado y espera lectura de boton.
    GPIO.output(ledS,GPIO.HIGH) #enciende led de estado
    

    time.sleep(0.5)#dejar que actue la raspberry
    while (flag == 0):
        print ("Incializando Arduino")
        #b1S=GPIO.input(b1)
        #time.sleep(0.5)
        ardR=ardS.readline()
        ardR1 = ardR.rstrip('\r\n')
        #print(type(ardR1))
        if (ardR1 == '1'): #se sale si presionas b1
            flag=1
            break
    print("Elegir modo de operacion")
    GPIO.output(ledS,GPIO.LOW)
    #GPIO.output(nmos,GPIO.LOW)

    time.sleep(0.5) #esperar por cualquier eventualidad
    b1S=GPIO.input(b1) #ALMACENO EL ESTADO inicial
    b2S=GPIO.input(b2) #ALMACENO EL ESTADO inicial

    if (b1S==1 and b2S == 1): #SIN  PRESIONAR ESTA EN BAJO
        print("No puede presionar ambos botones a la vez")
    else:
        while (b1S==0 and b2S == 0): #permanecera en el loop hasta que se presione un boton
            #print("elija el modo")
            GPIO.output(ledS,GPIO.HIGH)
            b1S=GPIO.input(b1) #ALMACENO EL ESTADO ACTUAL
            b2S=GPIO.input(b2) #ALMACENO EL ESTADO ACTUAL
            time.sleep(0.25)
            if (b1S == 1):
                b1F=1 #bandera para modo maze
                GPIO.output(ledS,GPIO.LOW)
                #print(b1F)
                break
            #time.sleep(0.25)
            if (b2S==1):
                b2F=2 #bandera para modo manual
                GPIO.output(ledS,GPIO.LOW)
                #print(b2F)
                break #SALE DEL WHILE SI ALGUNO DE LOS DOS ES PRESIONADO
            GPIO.output(ledS,GPIO.LOW)

####################MODO MAZE###################################
    while b1F == 1: #Modo maze activo

        # control de modo debug:
        while (flag == 1): #bandera flag reutilizada
            print(" Modo Maze-->desea activar modo maze con debug o !debug")
            b1S = GPIO.input(b1)  # lectura de estado de botones
            b2S = GPIO.input(b2)
            time.sleep(0.25)
            if (b1S == 1):  # debug
                flag = 0
                debug=1
            ######TCP INICIALIZACION#####
                try:
                    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                except socket.error:
                    print('Error Socket')
                    print("Modulos Desactivados")
                    ardS.write(b'E')
                    ardS.write(b's')
                    time.sleep(2)
                    #os.excel("restart.sh","")
                    sys.exit()

                s.bind((HOST, PORT))
                s.listen(3)  # Maximo tres clientes
                print("Creando Socket_Modo_Maze")
                conn, addr = s.accept() #se queda esperando un cliente
                print('Connected by', addr)
            ######TCP INICIALIZACION#####
            ####Thread Inicializacion####
                print_lock.acquire()
                start_new_thread(threadedMaze, (q,))
                print("Thread iniciado Modo Maze debug")
            ####Thread Inicializacion####
                break
            if (b2S == 1):  # !debug
                ####Thread Inicializacion####
                print_lock.acquire()
                start_new_thread(threadedMaze, (q,))
                print("Thread iniciado Modo Maze debug!")
                ####Thread Inicializacion####
                flag = 0
                debug =0
                break

        if enaS == 1:  # una sola ejecucion
            enaS = 0
            for i in range(3):  # parpadeo de led de estado 3 veces indica modo maze.
                GPIO.output(ledS, GPIO.HIGH)
                time.sleep(1)
                GPIO.output(ledS, GPIO.LOW)

        ######ARDUINO ACTIVACION DE SISTEMA#####
        if (ena_Sensor == 0):
            ena_Sensor = 1
            ardS.write(b'R')  # Activacion de los sensores
            time.sleep(0.080)
            ardS.write(b's')  # fin de mensaje
            print("UART establecido")
        ######ARDUINO ACTIVACION DE SISTEMA#####
        
        ###Adquisicion de data y procesamiento###
        rpiard(1)
        ###Adquisicion de data y procesamiento###

        b1S=GPIO.input(b1)
        b2S=GPIO.input(b2)
        if (b1S == 1): #detencion del programa manualmente
            print ("Modulos Desactivados")
            if (debug):
                conn.sendall('0') #indica que el proceso de graficado debe acabar
                time.sleep(1)
                conn.close()
            ardS.write(b"E") #Desactiva los sensores
            ardS.write(b"s")
            time.sleep(2)
            GPIO.cleanup()
	    #os.excel("restart.sh","")
            sys.exit()
        elif (b2S == 1):
            print ("Modulos Desactivados")
            if (debug):
                conn.sendall('0') #indica que el procesdo de graficado debe acabar
                time.slee(1)
                conn.close()
            ardS.write(b"E") #Desactiva los sensores
            ardS.write(b"s")
            time.sleep(2)
            GPIO.cleanup()
            sys.exit()


 #########################MODO MANUAL###############################################
    while b2F == 2: #Modo manual activo
        if ena2S == 1:
            ena2S=0
            print("Modo manual Activo")
            for i in range(5): #parpadeo 5 veces modo manual
                GPIO.output(ledS, GPIO.HIGH)
                time.sleep(0.5)  # parpadeo indica modo manual
                GPIO.output(ledS, GPIO.LOW)
            ####conexion Bluetooth ######
            server_sock=bluetooth.BluetoothSocket( bluetooth.RFCOMM )
            port = 0x1101 #make sure you've set up an RFCOMM sdptool
            server_sock.bind(("",port))
            server_sock.listen(1)
            try:
                print("Esperando Conexion")
                c,address = server_sock.accept()
                print ("Conexion Bluetooth establecida ",address)
                print_lock.acquire()
                #client_sock.setblocking(0)
                start_new_thread(threaded, (c,))
                print("Thread iniciado")
            except Exception as e:
                print("No se ha podido establecer la conexion")
                server_sock.close()

            ####conexion TCP ######
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.bind((HOST, PORT))
            s.listen(5)
            print("Creando Socket_modo_Manual")
            conn, addr = s.accept()
            print('Connected by', addr) #se queda esperando un cliente
            ####conexion TCP ######
        ######ARDUINO ACTIVACION DE SISTEMA#####
        if (ena_Sensor == 0):
            ena_Sensor = 1
            ardS.write(b'R')  # Activacion de los sensores
            time.sleep(0.080)
            ardS.write(b's')  # fin de mensaje
            print("UART establecido")
        ######ARDUINO ACTIVACION DE SISTEMA#####
        rpiard(0) # separa datos, envia por tcp

        b1S=GPIO.input(b1)
        b2S=GPIO.input(b1)
        if (b1S == 1): #detencion del programa manualmente
            print ("Modulos Desactivados")
            conn.sendall('1') #indica que el procesdo de graficado debe acabar
            time.sleep(1)
            conn.close()
            ardS.write(b"E") #Desactiva los sensores
            ardS.write(b"s")
            time.sleep(2)
            GPIO.cleanup()
	    #os.excel("restart.sh","")
            sys.exit()
        elif (b2S == 1):
            print ("Modulos Desactivados")
            if (debug):
                conn.sendall('0') #indica que el procesdo de graficado debe acabar
                time.slee(1)
                conn.close()
            ardS.write(b"E") #Desactiva los sensores
            ardS.write(b"s")
            time.sleep(2)
            GPIO.cleanup()
            sys.exit()


except KeyboardInterrupt :
    print ("Modulos Desactivados")
    if (debug):
        conn.sendall(packet_toSend)
        time.slee(1)
        conn.close()
    ardS.write(b"E")
    ardS.write(b"s")
    time.sleep(2)
    GPIO.cleanup()
    sys.exit()
