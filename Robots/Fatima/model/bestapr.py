import openpyxl
from openpyxl import load_workbook
from scipy import stats
import numpy as np
import matplotlib.pyplot as plt
import math
#array encargado de almacenar data
deg=[0]*353
d=[0]*353
x=[0]*353
y=[0]*353
i=0
#cargar documento
wb= load_workbook('old/reg0-180.xlsx')
sheet= wb['reg0-45']
#iteracion en las celdas
for data in range(1,176): #En lugar de 41 iria el tamano del vector que almacena r
    theta= sheet.cell(row=data, column=1).value
    r= sheet.cell(row=data, column=2).value
    #if (r !=80):
    data=data-1
    rect_x=r*math.cos(math.radians(theta))
    rect_y=r*math.sin(math.radians(theta))
    if (rect_x>-70 or rect_x<70):
    #deg[data]=math.radians(int(theta))
        x[data]=rect_x
    #d[data]=int(r)
    if(rect_y<=30):
        y[data]=rect_y
    #print(deg[data],d[data])

#Calculo de regresion
#slope, intercept, r_value, p_value, std_err = stats.linregress(deg,d)
#print("pendiente:",slope,"corte en y",intercept,"r^2",(r_value**2)*100)
#theta_test=np.linspace(5,45)
#y = slope*theta_test+intercept
#plt.plot(y, theta_test, '-r')
#plt.plot(deg, d)
#slope, intercept, r_value, p_value, std_err = stats.linregress(deg,d)
#print("pendiente:",slope,"corte en y",intercept,"r^2",(r_value**2)*100)
#theta_test=np.linspace(5,45)
#y = slope*theta_test+intercept
#plt.plot(y, theta_test, '-r')
#plt.polar(deg, d, 'ro')
plt.scatter(x, y)
plt.title('Graficando coordenadas rectangulares')
plt.xlabel('Distancia', color='#1C2833')
plt.ylabel('Distancia', color='#1C2833')
plt.grid()
plt.show()
