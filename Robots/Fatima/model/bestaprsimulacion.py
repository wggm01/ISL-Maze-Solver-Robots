import openpyxl
from openpyxl import load_workbook
import socket
from time import sleep as delay
#array encargado de almacenar data
#cargar documento
wb= load_workbook('reg0-180.xlsx')
sheet= wb['reg0-45']
data=0
HOST= '192.168.0.247'
PORT= 6793 # Revisar contra el cliente
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.bind((HOST, PORT))
s.listen(3)  # Maximo tres clientes
print("Creando Socket")
conn, addr = s.accept() #se queda esperando un cliente
print('Connected by', addr)
#iteracion en las celdas
while(True): #En lugar de 41 iria el tamano del vector que almacena r
    data+=1
    if(data==3094):
        print('Bye Bye')
        break
    theta= sheet.cell(row=data, column=1).value
    r= sheet.cell(row=data, column=2).value
    data_to = str(theta)+','+str(r)+'s'
    data_toSend=data_to.encode('utf-8')
    packet_len=len(data_toSend)
    diff=8-packet_len
    packet_adj=data_to + 's'*diff
    packet_toSend=packet_adj.encode('utf-8')
    print("bytes a enviar",len(packet_toSend),"bytes no adj",packet_len,"diferencia",diff)
    try:
        conn.sendall(packet_toSend)
    except socket.error:
        print ('No se pudo enviar la informacion')
        conn.close()
    delay(0.0001)
    
